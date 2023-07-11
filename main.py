import os
from binance.client import Client
from datetime import datetime
import time
from openpyxl import Workbook
import telebot


# API ключи Binance
api_key = ('ghGCOpbb4frqMLiaHs3Sp0jJmHDjOFOa2UpCRUyKdqaL43VuObV3jgFlTPJhU8Eg')
api_secret = ('wwOjNBXLq2VyIKBw3rNMpodMk7kBT01GWkSrsXAyAboJ7Y2pDe6hvjqOfgwLXDo7')


# Инициализация клиента
client = Client(api_key, api_secret)


# Excel таблица для сбора данных
wb = Workbook()
ws = wb.active

# Название колонок для данных
ws['A1'] = 'Average'
ws['B1'] = 'Futures'
ws['C1'] = 'Percentage'
ws['D1'] = 'Time'


# Задаём торговую пару
ASSET = 'BTCUSDT'


# Получаем цену
def price(symbol):
    try:
        price = client.get_avg_price(symbol=symbol, requests_params={'timeout': 2})['price']
        return float(price)
    except Exception as e:
        print(e)


# Получаем цену фьючерса
def priceF(symbol):
    try:
        priceF = client.futures_symbol_ticker(symbol=symbol, requests_params={'timeout': 2})['price']
        return float(priceF)
    except Exception as e:
        print(e)


# Инициализация бота Телеграм
api_telegram = ('6378829204:AAHtPK8eBcpn0r6PAGlMdcgutLyq5IVJROI')
bot = telebot.TeleBot(api_telegram)


# Идентификация ID
ID = '1247655421'


# Функция отправки сообщений
def message(text):
    bot.send_message(ID, text)


# Функция отправки сигнала
def message_signal():
    message(f'Сигнал по торговой паре {ASSET}')


# Инициализируем время между циклами
TIME = 1


# Процент выше которого получаем сигнал
GROWTH_PERCENT = 0.053


# Основной рабочий цикл
while True:
    # Сравниваем цену фьючерса со средней ценой
    FIRST_PRICE = price(ASSET)
    PRICEF = priceF(ASSET)
    PERCENT = ((PRICEF - FIRST_PRICE) / FIRST_PRICE) * 100


    if PERCENT >= GROWTH_PERCENT:
        # Вывод результатов в консоль
        print(ASSET)
        print('LOOK')
        print(PERCENT)
        times = datetime.now().strftime("%H:%M:%S")
        print(times)

        # Добавляем данные в таблицу
        ws.append([FIRST_PRICE, PRICEF, PERCENT, times])
        wb.save("data.xlsx")

    time.sleep(TIME)